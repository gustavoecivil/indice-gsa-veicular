"""
Ingestao do preco medio semanal de combustiveis por UF, direto da fonte
oficial da ANP (gov.br/anp).

Por que direto da ANP e nao via Base dos Dados / BigQuery: o dataset
tratado "br_anp_precos_combustiveis" do basedosdados.org roda em cima do
BigQuery, e exige projeto/credencial de faturamento do Google Cloud que
nao temos disponivel neste ambiente. A propria ANP publica a mesma serie
historica semanal por estado, pronta, em um unico arquivo .xlsx publico
sem necessidade de autenticacao — entao usamos essa fonte direto em vez
de travar esperando credencial do BigQuery.

Fonte: "Levantamento de Precos de Combustiveis" (Serie Historica
Semanal por Estado, desde 30/12/2012):
    https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/
    precos/precos-revenda-e-de-distribuicao-combustiveis/shlp/semanal/
    semanal-estados-desde-2013.xlsx

O arquivo cobre gasolina comum/aditivada, etanol hidratado, oleo diesel
(comum e S10), GLP e GNV, com preco medio/minimo/maximo/desvio padrao de
revenda por semana e por UF, desde dezembro de 2012 ate a semana mais
recente pesquisada.

Por padrao so carrega os ultimos N anos (--anos-recentes, default 3) pra
manter a primeira rodada enxuta — o arquivo inteiro tem toda a serie
desde 2012 e pode ser carregado com --tudo quando fizer sentido expandir
o historico.

Grava em data/processed/indice_gsa.duckdb, tabela anp_precos_combustivel.

Uso:
    python scripts/ingestao/anp.py                    # ultimos 3 anos
    python scripts/ingestao/anp.py --anos-recentes 5   # ultimos 5 anos
    python scripts/ingestao/anp.py --tudo              # serie completa desde 2012
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import duckdb
import openpyxl
import requests

BASE_DIR = Path(__file__).resolve().parents[2]
DB_PATH = BASE_DIR / "data" / "processed" / "indice_gsa.duckdb"
XLSX_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/precos-revenda-e-de-distribuicao-combustiveis/shlp/semanal/"
    "semanal-estados-desde-2013.xlsx"
)
XLSX_PATH = BASE_DIR / "data" / "raw" / "anp_semanal_estados_desde_2013.xlsx"
SHEET_NAME = "ESTADOS - DESDE 30.12.2012"
HEADER_ROW = 18
FIRST_DATA_ROW = 19
REQUEST_TIMEOUT = 120

# De-para "ESTADO" (nome por extenso, como publicado pela ANP) -> sigla
# de UF. Mapeamento fixo e objetivo (nao e uma classificacao subjetiva),
# cobre as 26 UFs + Distrito Federal que aparecem na planilha.
UF_POR_ESTADO = {
    "ACRE": "AC",
    "ALAGOAS": "AL",
    "AMAPA": "AP",
    "AMAZONAS": "AM",
    "BAHIA": "BA",
    "CEARA": "CE",
    "DISTRITO FEDERAL": "DF",
    "ESPIRITO SANTO": "ES",
    "GOIAS": "GO",
    "MARANHAO": "MA",
    "MATO GROSSO": "MT",
    "MATO GROSSO DO SUL": "MS",
    "MINAS GERAIS": "MG",
    "PARA": "PA",
    "PARAIBA": "PB",
    "PARANA": "PR",
    "PERNAMBUCO": "PE",
    "PIAUI": "PI",
    "RIO DE JANEIRO": "RJ",
    "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS",
    "RONDONIA": "RO",
    "RORAIMA": "RR",
    "SANTA CATARINA": "SC",
    "SAO PAULO": "SP",
    "SERGIPE": "SE",
    "TOCANTINS": "TO",
}


def normalizar_unidade(unidade: str) -> str:
    """Uniformiza variantes de grafia da mesma unidade ao longo dos anos
    (ex: 'R$/13Kg' vs 'R$/13kg', 'R$/m3' vs 'R$/m³')."""
    limpo = unidade.strip().upper().replace("³", "3").replace("�", "3")
    return limpo


def numero_ou_none(valor):
    if valor is None or valor == "-":
        return None
    return float(valor)


def inteiro_ou_none(valor):
    if valor is None or valor == "-":
        return None
    return int(valor)


def baixar_planilha() -> Path:
    print(f"[info] baixando planilha da ANP de {XLSX_URL} ...")
    resposta = requests.get(XLSX_URL, timeout=REQUEST_TIMEOUT)
    resposta.raise_for_status()
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    XLSX_PATH.write_bytes(resposta.content)
    print(f"[info] planilha salva em {XLSX_PATH} ({len(resposta.content):,} bytes)")
    return XLSX_PATH


def get_connection() -> duckdb.DuckDBPyConnection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(str(DB_PATH))
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS anp_precos_combustivel (
            uf VARCHAR NOT NULL,
            regiao VARCHAR NOT NULL,
            tipo_combustivel VARCHAR NOT NULL,
            unidade_medida VARCHAR NOT NULL,
            semana_referencia DATE NOT NULL,
            semana_fim DATE NOT NULL,
            preco_medio DOUBLE NOT NULL,
            preco_minimo DOUBLE,
            preco_maximo DOUBLE,
            desvio_padrao DOUBLE,
            numero_postos_pesquisados INTEGER,
            atualizado_em TIMESTAMP NOT NULL,
            PRIMARY KEY (uf, tipo_combustivel, semana_referencia)
        )
        """
    )
    return con


def carregar_linhas(xlsx_path: Path, desde: date | None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[SHEET_NAME]

    cabecalho = next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)
    )
    colunas_esperadas_prefixo = ("DATA INICIAL", "DATA FINAL")
    cabecalho_normalizado = tuple(
        (c or "").strip().upper().replace("�", "") for c in cabecalho[:2]
    )
    if cabecalho_normalizado != colunas_esperadas_prefixo:
        raise ValueError(
            f"Layout da planilha mudou — esperava colunas iniciando com "
            f"{colunas_esperadas_prefixo} na linha {HEADER_ROW}, encontrei "
            f"{cabecalho[:2]!r}. Confira o arquivo antes de importar."
        )

    now = datetime.now(timezone.utc)
    linhas_ignoradas_uf_desconhecida = 0

    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        data_inicial = row[0]
        if data_inicial is None:
            continue

        semana_referencia = data_inicial.date()
        if desde is not None and semana_referencia < desde:
            continue

        estado = (row[3] or "").strip().upper()
        uf = UF_POR_ESTADO.get(estado)
        if uf is None:
            linhas_ignoradas_uf_desconhecida += 1
            continue

        yield (
            uf,
            (row[2] or "").strip().upper(),
            (row[4] or "").strip().upper(),
            normalizar_unidade(row[6] or ""),
            semana_referencia,
            row[1].date(),
            float(row[7]),
            numero_ou_none(row[9]),
            numero_ou_none(row[10]),
            numero_ou_none(row[8]),
            inteiro_ou_none(row[5]),
            now,
        )

    if linhas_ignoradas_uf_desconhecida:
        print(
            f"[aviso] {linhas_ignoradas_uf_desconhecida} linha(s) ignorada(s) por "
            "ter nome de estado nao reconhecido no mapeamento UF_POR_ESTADO — "
            "confira se a ANP adicionou alguma localidade nova."
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    grupo = parser.add_mutually_exclusive_group()
    grupo.add_argument(
        "--anos-recentes",
        type=int,
        default=3,
        help="Carrega so os ultimos N anos de historico (default: 3).",
    )
    grupo.add_argument(
        "--tudo",
        action="store_true",
        help="Carrega a serie completa, desde 30/12/2012.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    desde = None if args.tudo else date.today() - timedelta(days=365 * args.anos_recentes)

    xlsx_path = baixar_planilha()

    con = get_connection()
    try:
        antes = con.execute("SELECT COUNT(*) FROM anp_precos_combustivel").fetchone()[0]

        linhas = list(carregar_linhas(xlsx_path, desde))
        if not linhas:
            print("[erro] nenhuma linha carregada da planilha — abortando sem gravar.")
            sys.exit(1)

        con.executemany(
            """
            INSERT OR REPLACE INTO anp_precos_combustivel VALUES
            (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            linhas,
        )

        depois = con.execute("SELECT COUNT(*) FROM anp_precos_combustivel").fetchone()[0]
        ufs_distintas = con.execute(
            "SELECT COUNT(DISTINCT uf) FROM anp_precos_combustivel"
        ).fetchone()[0]
        combustiveis_distintos = con.execute(
            "SELECT DISTINCT tipo_combustivel FROM anp_precos_combustivel ORDER BY 1"
        ).fetchall()
        semana_min, semana_max = con.execute(
            "SELECT MIN(semana_referencia), MAX(semana_referencia) "
            "FROM anp_precos_combustivel"
        ).fetchone()

        print("\n===== Resumo da ingestao ANP =====")
        print(f"Escopo desta rodada: "
              f"{'serie completa' if desde is None else f'desde {desde.isoformat()}'}")
        print(f"Linhas lidas da planilha nesta rodada: {len(linhas)}")
        print(f"anp_precos_combustivel antes:  {antes}")
        print(f"anp_precos_combustivel depois: {depois}")
        print(f"UFs distintas:      {ufs_distintas}")
        print(f"Combustiveis:       {[c[0] for c in combustiveis_distintos]}")
        print(f"Periodo no banco:   {semana_min} a {semana_max}")
    finally:
        con.close()


if __name__ == "__main__":
    main()
